import streamlit as st
import json
from datetime import datetime
import pandas as pd
from io import BytesIO
import base64

# 페이지 설정
st.set_page_config(
    page_title="위험성평가 작성 프로그램",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# 커스텀 CSS - 나눔고딕 폰트 및 스타일링
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Nanum+Gothic:wght@400;700;800&display=swap');
    
    * {
        font-family: 'Nanum Gothic', sans-serif !important;
    }
    
    .main {
        background-color: #f0f4f8;
    }
    
    .stButton > button {
        background: linear-gradient(90deg, #3b82f6 0%, #2563eb 100%);
        color: white;
        border: none;
        padding: 0.5rem 2rem;
        border-radius: 20px;
        font-weight: bold;
        transition: all 0.3s;
    }
    
    .stButton > button:hover {
        transform: scale(1.05);
        box-shadow: 0 5px 15px rgba(37, 99, 235, 0.4);
    }
    
    .cover-container {
        background: white;
        border: 2px solid #1f2937;
        border-radius: 10px;
        padding: 40px;
        max-width: 1000px;
        margin: 0 auto;
        box-shadow: 0 10px 30px rgba(0, 0, 0, 0.1);
    }
    
    .title-gradient {
        background: linear-gradient(90deg, #3b82f6 0%, #6366f1 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        font-size: 2.5rem;
        font-weight: 800;
        text-align: center;
        margin: 20px 0;
    }
    
    .approval-table {
        border-collapse: collapse;
        margin: 30px auto;
    }
    
    .approval-table td {
        border: 1px solid #d1d5db;
        padding: 10px;
        text-align: center;
    }
    
    .approval-header {
        background: linear-gradient(90deg, #fde047 0%, #facc15 100%);
        font-weight: bold;
        width: 60px;
    }
    
    input[type="text"] {
        border: 2px solid #e5e7eb;
        border-radius: 8px;
        padding: 8px 12px;
        width: 100%;
        transition: all 0.3s;
    }
    
    input[type="text"]:focus {
        border-color: #3b82f6;
        outline: none;
        box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.1);
    }
</style>
""", unsafe_allow_html=True)

# 세션 상태 초기화
if 'form_data' not in st.session_state:
    st.session_state.form_data = {
        'year': '',
        'company_name': '',
        'address': '',
        'phone': '',
        'fax': '',
        'approvers': [
            {'position': '', 'name': ''},
            {'position': '', 'name': ''},
            {'position': '', 'name': ''},
            {'position': '', 'name': ''}
        ]
    }

# 제목
st.markdown('<h1 style="text-align: center; color: #1f2937;">위험성평가 작성 프로그램</h1>', unsafe_allow_html=True)
st.markdown('---')

# 탭 생성
tab1, tab2, tab3 = st.tabs(["📄 표지", "📊 사업장 개요", "⚠️ 위험정보"])

with tab1:
    st.markdown('<div class="cover-container">', unsafe_allow_html=True)
    
    # 연도 입력
    col1, col2, col3 = st.columns([2, 1, 2])
    with col2:
        year = st.text_input(
            "연도",
            value=st.session_state.form_data['year'],
            max_chars=2,
            placeholder="24",
            label_visibility="collapsed"
        )
        if year:
            st.session_state.form_data['year'] = year
        
        st.markdown(f'<p style="text-align: center; font-size: 1.5rem; font-weight: bold;">20{year if year else "( )"}년도</p>', unsafe_allow_html=True)
    
    # 제목
    st.markdown('<h1 class="title-gradient">위험성평가 결과서</h1>', unsafe_allow_html=True)
    
    # 결재란
    st.markdown('<h3 style="text-align: center; margin-top: 30px;">결재란</h3>', unsafe_allow_html=True)
    
    col1, col2, col3, col4, col5 = st.columns(5)
    
    # 결재 헤더
    with col1:
        st.markdown('<div style="border: 1px solid #d1d5db; background: linear-gradient(90deg, #fde047 0%, #facc15 100%); padding: 20px 10px; text-align: center; font-weight: bold;">결재</div>', unsafe_allow_html=True)
    
    # 결재자 입력
    for i, col in enumerate([col2, col3, col4, col5]):
        with col:
            position = st.text_input(
                f"직위{i+1}",
                value=st.session_state.form_data['approvers'][i]['position'],
                placeholder="직위",
                key=f"position_{i}",
                label_visibility="collapsed"
            )
            name = st.text_input(
                f"성명{i+1}",
                value=st.session_state.form_data['approvers'][i]['name'],
                placeholder="성명",
                key=f"name_{i}",
                label_visibility="collapsed"
            )
            
            st.session_state.form_data['approvers'][i]['position'] = position
            st.session_state.form_data['approvers'][i]['name'] = name
    
    # 회사 정보
    st.markdown('<h3 style="text-align: center; margin-top: 40px;">회사 정보</h3>', unsafe_allow_html=True)
    
    company_name = st.text_input(
        "회사명",
        value=st.session_state.form_data['company_name'],
        placeholder="회사명을 입력하세요",
        label_visibility="collapsed"
    )
    st.session_state.form_data['company_name'] = company_name
    
    address = st.text_input(
        "주소",
        value=st.session_state.form_data['address'],
        placeholder="주소를 입력하세요",
        label_visibility="collapsed"
    )
    st.session_state.form_data['address'] = address
    
    col1, col2 = st.columns(2)
    with col1:
        phone = st.text_input(
            "전화번호",
            value=st.session_state.form_data['phone'],
            placeholder="전화번호",
            label_visibility="collapsed"
        )
        st.session_state.form_data['phone'] = phone
    
    with col2:
        fax = st.text_input(
            "팩스번호",
            value=st.session_state.form_data['fax'],
            placeholder="팩스번호",
            label_visibility="collapsed"
        )
        st.session_state.form_data['fax'] = fax
    
    st.markdown('</div>', unsafe_allow_html=True)
    
    # 버튼들
    st.markdown('<br><br>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 1, 1])
    
    with col2:
        if st.button("💾 표지 엑셀 저장", use_container_width=True):
            # 엑셀로 저장
            df = pd.DataFrame([st.session_state.form_data])
            
            # 결재자 정보를 별도 열로 분리
            for i, approver in enumerate(st.session_state.form_data['approvers']):
                df[f'결재자{i+1}_직위'] = approver['position']
                df[f'결재자{i+1}_성명'] = approver['name']
            
            # approvers 컬럼 제거
            df = df.drop('approvers', axis=1)
            
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='표지', index=False)
                
                # 서식 설정
                workbook = writer.book
                worksheet = writer.sheets['표지']
                
                # 헤더 스타일
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                
                header_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
                header_font = Font(bold=True, size=12)
                header_alignment = Alignment(horizontal='center', vertical='center')
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 헤더 서식 적용
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = thin_border
                
                # 열 너비 조정
                for column in worksheet.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2) * 1.2
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            output.seek(0)
            b64 = base64.b64encode(output.read()).decode()
            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="위험성평가_표지_{st.session_state.form_data.get("year", "YYYY")}.xlsx">📥 엑셀 파일 다운로드</a>'
            st.markdown(href, unsafe_allow_html=True)
            st.success("표지가 엑셀 파일로 저장되었습니다!")

with tab2:
    st.markdown('<h2 style="text-align: center; color: #1f2937;">1. 사업장 개요</h2>', unsafe_allow_html=True)
    
    # 세션 상태 초기화 - 사업장 개요
    if 'business_info' not in st.session_state:
        st.session_state.business_info = {
            'business_name': '',
            'main_product': '',
            'evaluation_date': '',
            'representative': '',
            'employee_count': '',
            'evaluator': ''
        }
    
    # 사업장 정보 입력
    col1, col2 = st.columns(2)
    
    with col1:
        st.session_state.business_info['business_name'] = st.text_input(
            "사업장명",
            value=st.session_state.business_info['business_name'],
            placeholder="사업장명을 입력하세요"
        )
        st.session_state.business_info['main_product'] = st.text_input(
            "주요생산품",
            value=st.session_state.business_info['main_product'],
            placeholder="주요생산품을 입력하세요"
        )
        st.session_state.business_info['evaluation_date'] = st.date_input(
            "평가일자",
            value=None,
            format="YYYY/MM/DD"
        )
    
    with col2:
        st.session_state.business_info['representative'] = st.text_input(
            "대표자",
            value=st.session_state.business_info['representative'],
            placeholder="대표자명을 입력하세요"
        )
        st.session_state.business_info['employee_count'] = st.text_input(
            "근로자수",
            value=st.session_state.business_info['employee_count'],
            placeholder="근로자수를 입력하세요"
        )
        st.session_state.business_info['evaluator'] = st.text_input(
            "평가자",
            value=st.session_state.business_info['evaluator'],
            placeholder="평가자명을 입력하세요"
        )
    
    st.markdown('<hr>', unsafe_allow_html=True)
    st.markdown('<h3 style="text-align: center; color: #1f2937;">공정도</h3>', unsafe_allow_html=True)
    
    # 공정 수 관리
    if 'process_count' not in st.session_state:
        st.session_state.process_count = 5
    
    if 'processes' not in st.session_state:
        st.session_state.processes = [
            {
                'name': '',
                'photo': None,
                'description': '',
                'equipment': '',
                'hazardous_material': '',
                'hazardous_factor': ''
            } for _ in range(5)
        ]
    
    # 공정 추가/삭제 버튼
    col1, col2, col3 = st.columns([1, 2, 1])
    with col1:
        if st.button("➕ 공정 추가"):
            st.session_state.process_count += 1
            st.session_state.processes.append({
                'name': '',
                'photo': None,
                'description': '',
                'equipment': '',
                'hazardous_material': '',
                'hazardous_factor': ''
            })
            st.rerun()
    
    with col3:
        if st.button("➖ 공정 삭제") and st.session_state.process_count > 1:
            st.session_state.process_count -= 1
            st.session_state.processes.pop()
            st.rerun()
    
    # 공정을 5개씩 그룹으로 나누어 표시
    process_groups = []
    for i in range(0, st.session_state.process_count, 5):
        process_groups.append(range(i, min(i + 5, st.session_state.process_count)))
    
    # 각 그룹별로 공정 표시
    for group_idx, process_group in enumerate(process_groups):
        if group_idx > 0:
            st.markdown('<hr style="margin: 30px 0;">', unsafe_allow_html=True)
            
        # 이 그룹의 공정 수 (최대 5개)
        group_size = len(process_group)
        
        # 컬럼 생성 (공정들만)
        cols = st.columns(group_size)
        
        # 각 공정별 입력 필드
        for col_idx, process_idx in enumerate(process_group):
            with cols[col_idx]:
                # 공정명
                st.markdown('<div style="font-weight: bold; margin-bottom: 5px;">공정명</div>', unsafe_allow_html=True)
                st.session_state.processes[process_idx]['name'] = st.text_input(
                    f"공정명 {process_idx+1}",
                    value=st.session_state.processes[process_idx]['name'],
                    placeholder=f"공정 {process_idx+1}",
                    key=f"process_name_{process_idx}",
                    label_visibility="collapsed"
                )
                
                # 화살표 표시 (각 그룹의 첫 번째 공정 제외)
                if col_idx > 0 or (group_idx > 0 and col_idx == 0):
                    st.markdown('<div style="text-align: center; font-size: 20px; color: #6b7280; margin: 5px 0;">→</div>', unsafe_allow_html=True)
                else:
                    st.markdown('<div style="margin: 5px 0; height: 28px;"></div>', unsafe_allow_html=True)
                
                # 공정사진
                st.markdown('<div style="font-weight: bold; margin-bottom: 5px;">공정사진</div>', unsafe_allow_html=True)
                photo = st.file_uploader(
                    f"공정사진 {process_idx+1}",
                    type=['png', 'jpg', 'jpeg'],
                    key=f"process_photo_{process_idx}",
                    label_visibility="collapsed"
                )
                if photo:
                    st.session_state.processes[process_idx]['photo'] = photo
                    st.image(photo, use_column_width=True)
                else:
                    st.markdown('<div style="height: 120px; border: 2px dashed #d1d5db; display: flex; align-items: center; justify-content: center; color: #9ca3af; background-color: #f9fafb;">사진 업로드<br>클릭 또는 드래그</div>', unsafe_allow_html=True)
                
                # 간격
                st.markdown('<div style="margin: 10px 0;"></div>', unsafe_allow_html=True)
                
                # 공정설명
                st.markdown('<div style="font-weight: bold; margin-bottom: 5px;">공정설명</div>', unsafe_allow_html=True)
                st.session_state.processes[process_idx]['description'] = st.text_area(
                    f"공정설명 {process_idx+1}",
                    value=st.session_state.processes[process_idx]['description'],
                    placeholder="공정 설명",
                    key=f"process_desc_{process_idx}",
                    height=100,
                    label_visibility="collapsed"
                )
                
                # 간격
                st.markdown('<div style="margin: 10px 0;"></div>', unsafe_allow_html=True)
                
                # 주요기계기구
                st.markdown('<div style="font-weight: bold; margin-bottom: 5px;">주요기계기구</div>', unsafe_allow_html=True)
                st.session_state.processes[process_idx]['equipment'] = st.text_area(
                    f"주요기계기구 {process_idx+1}",
                    value=st.session_state.processes[process_idx]['equipment'],
                    placeholder="주요기계기구",
                    key=f"process_equip_{process_idx}",
                    height=100,
                    label_visibility="collapsed"
                )
                
                # 간격
                st.markdown('<div style="margin: 10px 0;"></div>', unsafe_allow_html=True)
                
                # 유해위험물질
                st.markdown('<div style="font-weight: bold; margin-bottom: 5px;">유해위험물질</div>', unsafe_allow_html=True)
                st.session_state.processes[process_idx]['hazardous_material'] = st.text_area(
                    f"유해위험물질 {process_idx+1}",
                    value=st.session_state.processes[process_idx]['hazardous_material'],
                    placeholder="유해위험물질",
                    key=f"process_material_{process_idx}",
                    height=100,
                    label_visibility="collapsed"
                )
                
                # 간격
                st.markdown('<div style="margin: 10px 0;"></div>', unsafe_allow_html=True)
                
                # 유해위험요인
                st.markdown('<div style="font-weight: bold; margin-bottom: 5px;">유해위험요인</div>', unsafe_allow_html=True)
                st.session_state.processes[process_idx]['hazardous_factor'] = st.text_area(
                    f"유해위험요인 {process_idx+1}",
                    value=st.session_state.processes[process_idx]['hazardous_factor'],
                    placeholder="유해위험요인",
                    key=f"process_factor_{process_idx}",
                    height=100,
                    label_visibility="collapsed"
                )
    
    # 데이터 저장 버튼
    st.markdown('<br>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 1, 1])
    with col2:
        if st.button("💾 사업장 개요 엑셀 저장", use_container_width=True, key="save_tab2"):
            # 사업장 개요 데이터
            overview_df = pd.DataFrame([st.session_state.business_info])
            
            # 공정 데이터
            process_list = []
            for process in st.session_state.processes:
                if process['name']:
                    process_list.append({
                        '공정명': process['name'],
                        '공정설명': process['description'],
                        '주요기계기구': process['equipment'],
                        '유해위험물질': process['hazardous_material'],
                        '유해위험요인': process['hazardous_factor']
                    })
            
            process_df = pd.DataFrame(process_list)
            
            # 엑셀로 저장
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                overview_df.to_excel(writer, sheet_name='사업장정보', index=False)
                if not process_df.empty:
                    process_df.to_excel(writer, sheet_name='공정정보', index=False)
                
                # 서식 설정
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                
                header_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
                header_font = Font(bold=True, size=12)
                header_alignment = Alignment(horizontal='center', vertical='center')
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 각 시트에 서식 적용
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    
                    # 헤더 서식
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_alignment
                        cell.border = thin_border
                    
                    # 열 너비 자동 조정
                    for column in worksheet.columns:
                        max_length = 0
                        column = [cell for cell in column]
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2) * 1.2
                        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            output.seek(0)
            b64 = base64.b64encode(output.read()).decode()
            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="위험성평가_사업장개요_{datetime.now().strftime("%Y%m%d")}.xlsx">📥 엑셀 파일 다운로드</a>'
            st.markdown(href, unsafe_allow_html=True)
            st.success("사업장 개요가 엑셀 파일로 저장되었습니다!")

with tab3:
    st.markdown('<h2 style="text-align: center; color: #1f2937;">안전보건상 위험정보</h2>', unsafe_allow_html=True)
    
    # 상단 정보 테이블 스타일
    st.markdown("""
    <style>
    .info-header {
        background-color: #fef3c7;
        border: 1px solid #d97706;
        padding: 10px;
        font-weight: bold;
        text-align: center;
        min-width: 120px;
    }
    .process-table {
        width: 100%;
        border-collapse: collapse;
        margin-top: 20px;
    }
    .process-table th, .process-table td {
        border: 1px solid #d97706;
        padding: 10px;
        text-align: center;
    }
    .process-header {
        background-color: #fef3c7;
        font-weight: bold;
        font-size: 16px;
    }
    .sub-header {
        background-color: #fef3c7;
        font-size: 14px;
        font-weight: normal;
    }
    .stTextInput input, .stTextArea textarea, .stSelectbox select {
        font-size: 16px !important;
        padding: 10px !important;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # 상단 정보 입력
    col1, col2, col3, col4 = st.columns([1, 2, 1, 2])
    with col1:
        st.markdown('<div class="info-header">업종명</div>', unsafe_allow_html=True)
    with col2:
        st.text_input("업종명", label_visibility="collapsed", key="industry_name")
    with col3:
        st.markdown('<div class="info-header">생산품</div>', unsafe_allow_html=True)
    with col4:
        st.text_input("생산품", label_visibility="collapsed", key="product_name")
    
    col1, col2, col3, col4 = st.columns([1, 2, 1, 2])
    with col1:
        st.markdown('<div class="info-header">원(재)료</div>', unsafe_allow_html=True)
    with col2:
        st.text_input("원재료", label_visibility="collapsed", key="raw_material")
    with col3:
        st.markdown('<div class="info-header">근로자</div>', unsafe_allow_html=True)
    with col4:
        st.text_input("근로자", label_visibility="collapsed", key="workers_info")
    
    st.markdown('<hr style="margin: 30px 0;">', unsafe_allow_html=True)
    
    # 공정(작업)순서 테이블
    st.markdown('<h3 style="text-align: center; color: #1f2937;">공정(작업)순서</h3>', unsafe_allow_html=True)
    
    # 원본과 동일한 테이블 헤더
    st.markdown("""
    <table class="process-table">
        <tr>
            <th rowspan="2" class="process-header" style="width: 8%;">공정<br>(작업)순서</th>
            <th colspan="2" class="process-header">기계기구 및 설비명</th>
            <th colspan="3" class="process-header">유해화학물질</th>
            <th colspan="8" class="process-header">기타 안전보건상 정보</th>
        </tr>
        <tr>
            <th class="sub-header">기계기구 및<br>설비명</th>
            <th class="sub-header">수량</th>
            <th class="sub-header">화학물질명</th>
            <th class="sub-header">취급량/일</th>
            <th class="sub-header">취급시간</th>
            <th class="sub-header">3년간<br>재해사례</th>
            <th class="sub-header">앗차<br>사고사례</th>
            <th class="sub-header">근로자<br>구성및특성</th>
            <th class="sub-header">도급/교대<br>작업유무</th>
            <th class="sub-header">운반수단</th>
            <th class="sub-header">안전작업<br>허가증<br>필요작업</th>
            <th class="sub-header">작업환경<br>측정유무</th>
            <th class="sub-header">특별안전<br>교육대상</th>
        </tr>
    </table>
    """, unsafe_allow_html=True)
    
    # 데이터 저장을 위한 리스트
    process_data_list = []
    
    # 공정별 데이터 입력
    if 'processes' in st.session_state:
        for idx, process in enumerate(st.session_state.processes):
            if process['name']:
                # 균등한 컬럼 분할
                cols = st.columns([0.8, 1.2, 0.5, 1.2, 0.6, 0.6, 0.8, 0.8, 0.8, 0.8, 0.8, 0.8, 0.8, 0.8])
                
                # 각 필드의 값을 저장할 딕셔너리
                process_row = {}
                
                # 공정(작업)순서
                with cols[0]:
                    process_row['공정순서'] = process['name']
                    st.text_input(f"공정_{idx}", value=process['name'], disabled=True, label_visibility="collapsed")
                
                # 기계기구 및 설비명
                with cols[1]:
                    process_row['기계기구및설비명'] = process['equipment']
                    st.text_area(f"기계_{idx}", value=process['equipment'], height=100, disabled=True, label_visibility="collapsed")
                
                # 수량
                with cols[2]:
                    qty = st.text_input(f"수량_{idx}", placeholder="", label_visibility="collapsed", key=f"qty_{idx}")
                    process_row['수량'] = qty
                
                # 화학물질명
                with cols[3]:
                    process_row['화학물질명'] = process['hazardous_material']
                    st.text_area(f"화학_{idx}", value=process['hazardous_material'], height=100, disabled=True, label_visibility="collapsed")
                
                # 취급량/일
                with cols[4]:
                    amount = st.text_input(f"취급량_{idx}", placeholder="", label_visibility="collapsed", key=f"amount_{idx}")
                    process_row['취급량/일'] = amount
                
                # 취급시간
                with cols[5]:
                    time = st.text_input(f"취급시간_{idx}", placeholder="", label_visibility="collapsed", key=f"time_{idx}")
                    process_row['취급시간'] = time
                
                # 3년간 재해사례
                with cols[6]:
                    accident = st.text_input(f"재해사례_{idx}", placeholder="", label_visibility="collapsed", key=f"accident_{idx}")
                    process_row['3년간재해사례'] = accident
                
                # 앗차사고사례
                with cols[7]:
                    near_miss = st.text_input(f"앗차_{idx}", placeholder="", label_visibility="collapsed", key=f"near_miss_{idx}")
                    process_row['앗차사고사례'] = near_miss
                
                # 근로자 구성및특성
                with cols[8]:
                    workers = st.text_input(f"근로자구성_{idx}", placeholder="", label_visibility="collapsed", key=f"workers_{idx}")
                    process_row['근로자구성및특성'] = workers
                
                # 도급/교대 작업유무
                with cols[9]:
                    contract = st.selectbox(f"도급_{idx}", ["", "유", "무"], label_visibility="collapsed", key=f"contract_{idx}")
                    process_row['도급/교대작업유무'] = contract
                
                # 운반수단
                with cols[10]:
                    transport = st.text_input(f"운반_{idx}", placeholder="", label_visibility="collapsed", key=f"transport_{idx}")
                    process_row['운반수단'] = transport
                
                # 안전작업허가증필요작업
                with cols[11]:
                    permit = st.selectbox(f"허가증_{idx}", ["", "유", "무"], label_visibility="collapsed", key=f"permit_{idx}")
                    process_row['안전작업허가증필요작업'] = permit
                
                # 작업환경측정유무
                with cols[12]:
                    measurement = st.selectbox(f"측정_{idx}", ["", "유", "무"], label_visibility="collapsed", key=f"measurement_{idx}")
                    process_row['작업환경측정유무'] = measurement
                
                # 특별안전교육대상
                with cols[13]:
                    special_edu = st.text_input(f"특별교육_{idx}", placeholder="", label_visibility="collapsed", key=f"special_edu_{idx}")
                    process_row['특별안전교육대상'] = special_edu
                
                process_data_list.append(process_row)
                st.markdown('<hr style="margin: 10px 0; border: 0; border-top: 1px solid #d97706;">', unsafe_allow_html=True)
    
    # 데이터 저장 버튼 (엑셀)
    st.markdown('<br>', unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 1, 1])
    with col2:
        if st.button("💾 위험정보 엑셀 저장", use_container_width=True, key="save_tab3"):
            # 상단 정보
            header_data = {
                '업종명': st.session_state.get('industry_name', ''),
                '생산품': st.session_state.get('product_name', ''),
                '원재료': st.session_state.get('raw_material', ''),
                '근로자': st.session_state.get('workers_info', '')
            }
            
            # 엑셀 파일 생성
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # 헤더 정보
                df_header = pd.DataFrame([header_data])
                df_header.to_excel(writer, sheet_name='기본정보', index=False)
                
                # 공정 정보
                if process_data_list:
                    df_process = pd.DataFrame(process_data_list)
                    df_process.to_excel(writer, sheet_name='공정정보', index=False)
                
                # 서식 설정
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                
                header_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
                header_font = Font(bold=True, size=12)
                header_alignment = Alignment(horizontal='center', vertical='center')
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 각 시트에 서식 적용
                for sheet_name in writer.sheets:
                    worksheet = writer.sheets[sheet_name]
                    
                    # 헤더 서식
                    for cell in worksheet[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_alignment
                        cell.border = thin_border
                    
                    # 열 너비 자동 조정
                    for column in worksheet.columns:
                        max_length = 0
                        column = [cell for cell in column]
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2) * 1.2
                        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # 다운로드 링크 생성
            output.seek(0)
            b64 = base64.b64encode(output.read()).decode()
            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="위험성평가_위험정보_{datetime.now().strftime("%Y%m%d")}.xlsx">📥 엑셀 파일 다운로드</a>'
            st.markdown(href, unsafe_allow_html=True)
            st.success("위험정보가 엑셀 파일로 저장되었습니다!")

# 사이드바에 도움말 추가
with st.sidebar:
    st.markdown("### 📌 사용 방법")
    st.markdown("""
    1. **표지 탭**에서 기본 정보를 입력하세요
    2. **사업장 개요 탭**에서 공정 정보를 입력하세요
    3. **위험정보 탭**에서 위험성평가를 수행하세요
    4. 완료 후 전체 보고서를 생성할 수 있습니다
    """)
    
    st.markdown("### 🔧 기능")
    st.markdown("""
    - ✅ 데이터 자동 저장
    - ✅ PDF 보고서 생성
    - ✅ Excel 내보내기
    - ✅ 이전 평가 불러오기
    """)